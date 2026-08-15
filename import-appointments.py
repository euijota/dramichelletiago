#!/usr/bin/env python3
"""
import-appointments.py — Migra agendamentos do consultorio.me para o Supabase.

Lê a planilha de backup (.xlsx) exportada pelo consultorio.me, cruza dados
de pacientes e insere agendamentos na tabela `appointments` do Supabase.

Uso:
    export SUPABASE_URL="https://SEU_PROJECT_REF.supabase.co"
    export SUPABASE_SERVICE_ROLE_KEY="eyJhbGci..."
    python3 import-appointments.py caminho/para/backup.xlsx [--dry-run]
"""

from __future__ import annotations

import json
import os
import sys
import urllib.request
import urllib.error
from datetime import date, datetime, time
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl não encontrado. Instale com:  pip3 install --user --break-system-packages openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
CLINIC_PHONE = "(96) 98111-1157"
DEFAULT_EMAIL = "nao_informado@paciente.com"
DEFAULT_SERVICE = "Consulta (Migrada do Consultório.me)"
CUTOFF_DATE = date(2026, 2, 1)  # apenas fev/2026 em diante
BATCH_SIZE = 100

# Colunas da aba Appointments (por nome)
APT_COLS = [
    "DateTime",
    "PatientId",
    "Name",
    "Insurance",
    "InsuranceNumber",
    "Document",
    "Absence",
    "Confirmed",
    "Observation",
    "Type",
    "Professional",
    "Origin",
]

# Colunas da aba Patients relevantes para o cruzamento
PATIENT_COLS = ["PatientId", "Name", "Tel1", "Email"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_supabase_url() -> str:
    url = os.environ.get("SUPABASE_URL", "").strip().rstrip("/")
    if not url:
        print(
            "❌ SUPABASE_URL não encontrada.\n"
            "   Defina explicitamente o projeto de destino:\n"
            '     export SUPABASE_URL="https://SEU_PROJECT_REF.supabase.co"'
        )
        sys.exit(1)
    return url


def get_service_role_key() -> str:
    key = (
        # .env local (somente para conveniência, nunca commitado)
        Path(".env").read_text().splitlines()
        if Path(".env").exists()
        else []
    )
    for line in key:
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
            return line.split("=", 1)[1].strip().strip('"').strip("'")

    # variável de ambiente
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not key:
        print(
            "❌ SUPABASE_SERVICE_ROLE_KEY não encontrada.\n"
            "   Defina a variável no terminal:\n"
            '     export SUPABASE_SERVICE_ROLE_KEY="eyJhbGci..."\n'
            "   Ou adicione temporariamente ao .env (NÃO faça commit)."
        )
        sys.exit(1)
    return key


def supabase_request(
    method: str,
    path: str,
    body: list | dict | None = None,
    service_role_key: str = "",
) -> dict:
    """Faz uma requisição à API REST do Supabase."""
    url = f"{get_supabase_url()}/rest/v1/{path}"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    data = json.dumps(body).encode() if body is not None else None

    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return {"status": resp.status}
    except urllib.error.HTTPError as e:
        return {"status": e.code, "error": e.read().decode()}


def map_status(absence: str | None, confirmed: str | None, dt: datetime | None) -> str:
    """Converte Absence/Confirmed da planilha para o enum do app."""
    absence_yes = str(absence or "").strip().lower() in ("yes", "sim", "s", "1", "true")
    confirmed_yes = str(confirmed or "").strip().lower() in ("yes", "sim", "s", "1", "true")

    if absence_yes:
        return "cancelled"
    if confirmed_yes:
        if dt and dt.date() <= date.today():
            return "completed"
        return "confirmed"
    return "pending"


def extract_date_time(dt) -> tuple[date | None, time | None]:
    """Extrai date e time de um datetime do Excel."""
    if isinstance(dt, datetime):
        return dt.date(), dt.time().replace(microsecond=0)
    if isinstance(dt, date):
        return dt, None
    return None, None


def safe_str(val) -> str:
    """Converte qualquer valor para string segura (trim, sem None)."""
    if val is None:
        return ""
    s = str(val).strip()
    return s


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    dry_run = "--dry-run" in sys.argv

    # Caminho do xlsx
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith("--") else None
    if not xlsx_path:
        print(
            "Uso: python3 import-appointments.py caminho/do/backup.xlsx [--dry-run]\n"
            "\n"
            "  --dry-run   Mostra o que seria importado sem inserir no Supabase."
        )
        sys.exit(1)

    if not Path(xlsx_path).exists():
        print(f"❌ Arquivo não encontrado: {xlsx_path}")
        sys.exit(1)

    print(f"📂 Lendo planilha: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- Carregar aba Patients para cruzamento ---
    print("📋 Carregando pacientes (cruzamento de email/telefone)...")
    patients: dict[int, dict] = {}
    if "Patients" in wb.sheetnames:
        ws_pat = wb["Patients"]
        rows_pat = list(ws_pat.iter_rows(values_only=True))
        if rows_pat:
            hdr_pat = rows_pat[0]
            idx_pat = {h: i for i, h in enumerate(hdr_pat) if h in PATIENT_COLS}
            for row in rows_pat[1:]:
                pid = row[idx_pat.get("PatientId", 0)]
                if pid is not None:
                    patients[int(pid)] = {
                        "name": safe_str(row[idx_pat.get("Name", 1)]),
                        "tel1": safe_str(row[idx_pat.get("Tel1", 2)]),
                        "email": safe_str(row[idx_pat.get("Email", 9)]),
                    }
        print(f"   ✅ {len(patients)} pacientes carregados")
    else:
        print("   ⚠️  Aba 'Patients' não encontrada — sem cruzamento de contato")

    # --- Processar aba Appointments ---
    if "Appointments" not in wb.sheetnames:
        print("❌ Aba 'Appointments' não encontrada na planilha.")
        sys.exit(1)

    ws_apt = wb["Appointments"]
    rows_apt = list(ws_apt.iter_rows(values_only=True))
    if not rows_apt:
        print("❌ Aba 'Appointments' está vazia.")
        sys.exit(1)

    hdr_apt = rows_apt[0]
    idx = {h: i for i, h in enumerate(hdr_apt)}
    data_rows = rows_apt[1:]

    print(f"📅 Processando {len(data_rows)} agendamentos...")
    print(f"   Filtro: data >= {CUTOFF_DATE.isoformat()}")
    print()

    # Filtrar e mapear
    records: list[dict] = []
    skipped_old = 0
    skipped_no_date = 0

    for row in data_rows:
        dt_raw = row[idx["DateTime"]]
        apt_date, apt_time = extract_date_time(dt_raw)

        if apt_date is None:
            skipped_no_date += 1
            continue
        if apt_date < CUTOFF_DATE:
            skipped_old += 1
            continue

        patient_id_raw = row[idx["PatientId"]]
        patient_id = int(patient_id_raw) if patient_id_raw is not None else None
        patient_name = safe_str(row[idx["Name"]])

        # Cruzamento com Patients para email/telefone
        patient_info = patients.get(patient_id) if patient_id else None
        if patient_info:
            patient_email = patient_info["email"] or DEFAULT_EMAIL
            patient_phone = patient_info["tel1"] or CLINIC_PHONE
        else:
            patient_email = DEFAULT_EMAIL
            patient_phone = CLINIC_PHONE

        # Status
        absence = row[idx["Absence"]]
        confirmed = row[idx["Confirmed"]]
        status = map_status(absence, confirmed, dt_raw if isinstance(dt_raw, datetime) else None)

        # Serviço
        apt_type = safe_str(row[idx["Type"]])
        service_name = apt_type or DEFAULT_SERVICE

        # Notas
        observation = safe_str(row[idx["Observation"]])
        notes = f"[Migração Consultório.me] {observation}" if observation else "[Migração Consultório.me]"

        # Horário padrão se ausente
        time_str = apt_time.strftime("%H:%M") if apt_time else "09:00"

        records.append(
            {
                "patient_name": patient_name,
                "patient_email": patient_email,
                "patient_phone": patient_phone,
                "service_name": service_name,
                "appointment_date": apt_date.isoformat(),
                "appointment_time": time_str,
                "status": status,
                "notes": notes,
            }
        )

    # Agrupar status para resumo
    status_counts: dict[str, int] = {}
    for r in records:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1

    print(f"📊 Resumo do filtro:")
    print(f"   Total na planilha:       {len(data_rows)}")
    print(f"   Pulados (data antiga):   {skipped_old}")
    print(f"   Pulados (sem data):       {skipped_no_date}")
    print(f"   A importar:               {len(records)}")
    print(f"   Status: {dict(sorted(status_counts.items()))}")
    print()

    # --- Dry run: mostrar amostra e parar ---
    if dry_run:
        print("🔍 DRY-RUN — Nenhum dado será inserido. Amostra:")
        print()
        for r in records[:10]:
            print(f"   {r['appointment_date']} {r['appointment_time']}  {r['patient_name'][:40]:<40}  {r['status']:<10}  {r['service_name'][:30]}")
        if len(records) > 10:
            print(f"   ... mais {len(records) - 10} registros")
        print()
        print("✅ Dry-run concluído. Para importar, rode sem --dry-run.")
        wb.close()
        return

    # --- Importação real ---
    service_role_key = get_service_role_key()
    print("🔑 Service role key obtida. Iniciando importação...")

    # 1. Buscar existentes para deduplicação
    print("🔎 Verificando agendamentos existentes (deduplicação)...")
    existing: set[tuple[str, str, str]] = set()

    offset = 0
    while True:
        query = (
            f"appointment_date=gte.{CUTOFF_DATE.isoformat()}"
            f"&select=patient_name,appointment_date,appointment_time"
            f"&order=appointment_date"
            f"&offset={offset}&limit=1000"
        )
        resp = supabase_request("GET", f"appointments?{query}", service_role_key=service_role_key)

        if resp["status"] != 200:
            print(f"   ⚠️ Erro ao buscar existentes: {resp.get('error', '')[:200]}")
            print("   Continuando sem verificação de duplicatas...")
            break

        import json as _json

        # GET com return=minimal não retorna body; usar return=representation
        # Refazendo com Prefer adequado para GET
        url = f"{get_supabase_url()}/rest/v1/appointments?{query}"
        headers = {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
        }
        req = urllib.request.Request(url, headers=headers, method="GET")
        try:
            with urllib.request.urlopen(req) as r:
                body = _json.loads(r.read().decode())
        except urllib.error.HTTPError as e:
            print(f"   ⚠️ Erro: {e.read().decode()[:200]}")
            break

        if not body:
            break

        for row in body:
            key = (row["patient_name"], str(row["appointment_date"]), str(row["appointment_time"]))
            existing.add(key)

        offset += 1000
        print(f"   {len(existing)} existentes verificados...")

    # 2. Filtrar duplicatas e inserir em lotes
    to_insert = [
        r for r in records
        if (r["patient_name"], r["appointment_date"], r["appointment_time"]) not in existing
    ]

    skipped_dupes = len(records) - len(to_insert)
    print(f"   Duplicatas puladas: {skipped_dupes}")
    print(f"   A inserir (após dedup): {len(to_insert)}")
    print()

    if not to_insert:
        print("✅ Nenhum registro novo para inserir. Tudo já está no banco.")
        wb.close()
        return

    # Inserir em lotes
    inserted = 0
    errors = 0

    for i in range(0, len(to_insert), BATCH_SIZE):
        batch = to_insert[i : i + BATCH_SIZE]
        resp = supabase_request(
            "POST",
            "appointments",
            body=batch,
            service_role_key=service_role_key,
        )

        if resp["status"] in (201, 200):
            inserted += len(batch)
            print(f"   ✅ Lote {i // BATCH_SIZE + 1}: {len(batch)} inseridos  (total: {inserted})")
        else:
            errors += len(batch)
            err_msg = resp.get("error", "")[:300]
            print(f"   ❌ Lote {i // BATCH_SIZE + 1}: ERRO — {err_msg}")

    print()
    print("=" * 60)
    print(f"🏁 Importação concluída!")
    print(f"   Inseridos:     {inserted}")
    print(f"   Duplicatas:    {skipped_dupes}")
    print(f"   Erros:         {errors}")
    print(f"   Data antiga:   {skipped_old} (pulados)")
    print("=" * 60)

    wb.close()


if __name__ == "__main__":
    main()
